"""Write extracted documents to a downloadable workbook.

Sheet layout mirrors the AlgoDocs "Fields & Tables" panel: one sheet for the
document-level fields, one per table, plus a reconciliation sheet.
"""
from __future__ import annotations

from dataclasses import asdict
from io import BytesIO
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .northport import Document

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FORMAT = "#,##0.00"

FIELD_COLUMNS = [
    ("source_file", "Source File"),
    ("layout", "Layout"),
    ("bill_no", "Bill No"),
    ("invoice_date", "Invoice Date"),
    ("due_date", "Due Date"),
    ("account_no", "A/C No"),
    ("bill_to", "Bill To"),
    ("total_amount", "Total (RM)"),
    ("pages", "Pages"),
    ("detail_pages", "Detail Pages"),
    ("line_items", "Line Items"),
]

CHARGE_COLUMNS = [
    ("bill_no", "Bill No"),
    ("invoice_date", "Invoice Date"),
    ("due_date", "Due Date"),
    ("account_no", "A/C No"),
    ("reference_no", "Reference No"),
    ("location", "Location"),
    ("service_type", "Service Type"),
    ("bill_of_lading_no", "Bill Of Lading No"),
    ("purchase_order_no", "Purchase Order No"),
    ("sno", "SNO"),
    ("container_no", "Container No"),
    ("voyage", "Voyage"),
    ("size", "Size"),
    ("operator", "Oper"),
    ("status", "Status"),
    ("service_description", "Service Description"),
    ("description", "Description"),
    ("amount", "Amount (RM)"),
    ("sub_total", "Sub-Total (RM)"),
    ("source_file", "Source File"),
    ("page", "Page"),
]

TARIFF_COLUMNS = [
    ("bill_no", "Bill No"),
    ("invoice_date", "Invoice Date"),
    ("due_date", "Due Date"),
    ("account_no", "A/C No"),
    ("reference_no", "Reference No"),
    ("location", "Location"),
    ("ship_name", "Ship Name"),
    ("ship_id_voyage", "Ship ID / Voyage No"),
    ("loa", "LOA"),
    ("grt", "GRT"),
    ("group_ref", "Group Ref"),
    ("tariff_description", "Tariff Description"),
    ("tariff_code", "Tariff Code"),
    ("no_days", "No Days"),
    ("tariff_unit", "Tariff Unit"),
    ("rate", "Rate"),
    ("amount", "Amount (RM)"),
    ("sub_total", "Sub-Total (RM)"),
    ("source_file", "Source File"),
    ("page", "Page"),
]

SUMMARY_COLUMNS = [
    ("bill_no", "Bill No"),
    ("reference", "Reference"),
    ("location", "Location"),
    ("voyage_no", "Voyage No"),
    ("amount", "Amount (RM)"),
    ("source_file", "Source File"),
    ("page", "Page"),
]

# Columns whose header ends in (RM), plus Rate, get the money number format.
_MONEY_HEADERS = {"Amount (RM)", "Sub-Total (RM)", "Total (RM)", "Rate", "Variance"}


def _write_sheet(ws: Worksheet, columns: Sequence[tuple[str, str]], rows: Sequence[dict]) -> None:
    ws.append([label for _, label in columns])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([row.get(key) for key, _ in columns])

    widths: list[int] = []
    for index, (key, label) in enumerate(columns):
        longest = max(
            [len(label)] + [len(str(row.get(key) or "")) for row in rows] or [len(label)]
        )
        widths.append(min(max(longest + 2, 10), 55))
        letter = get_column_letter(index + 1)
        ws.column_dimensions[letter].width = widths[-1]
        if label in _MONEY_HEADERS:
            for cell in ws[letter][1:]:
                cell.number_format = MONEY_FORMAT
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def build_workbook(documents: Sequence[Document]) -> Workbook:
    wb = Workbook()

    fields = wb.active
    fields.title = "Fields"
    _write_sheet(fields, FIELD_COLUMNS, [asdict(d.fields) for d in documents])

    charges = [asdict(r) for d in documents for r in d.charge_rows]
    _write_sheet(wb.create_sheet("Table 1 - Charges"), CHARGE_COLUMNS, charges)

    tariffs = [asdict(r) for d in documents for r in d.tariff_rows]
    _write_sheet(wb.create_sheet("Table 2 - Tariff"), TARIFF_COLUMNS, tariffs)

    summaries = [asdict(r) for d in documents for r in d.summary_rows]
    _write_sheet(wb.create_sheet("Summary"), SUMMARY_COLUMNS, summaries)

    _write_sheet(wb.create_sheet("Validation"), VALIDATION_COLUMNS, _validation(documents))
    return wb


VALIDATION_COLUMNS = [
    ("source_file", "Source File"),
    ("bill_no", "Bill No"),
    ("check", "Check"),
    ("expected", "Expected"),
    ("actual", "Actual"),
    ("variance", "Variance"),
    ("status", "Status"),
    ("note", "Note"),
]


def _validation(documents: Sequence[Document]) -> list[dict]:
    rows: list[dict] = []
    for doc in documents:
        lines_total = sum(r.amount or 0 for r in doc.charge_rows)
        lines_total += sum(r.amount or 0 for r in doc.tariff_rows)
        summary_total = sum(r.amount or 0 for r in doc.summary_rows)
        bill_total = doc.fields.total_amount

        rows.append(_check(doc, "Summary rows vs Total (RM)", bill_total, summary_total,
                           "Every reference on the cover page adds up to the printed total."))
        rows.append(_check(doc, "Line items vs Total (RM)", bill_total, lines_total,
                           "Detail pages missing from the PDF show up here as a shortfall."))
        for warning in doc.warnings:
            rows.append({
                "source_file": doc.fields.source_file, "bill_no": doc.fields.bill_no,
                "check": "Page layout", "expected": None, "actual": None,
                "variance": None, "status": "REVIEW", "note": warning,
            })
    return rows


def _check(doc: Document, name: str, expected, actual, note: str) -> dict:
    variance = None if expected is None else round(actual - expected, 2)
    status = "OK" if variance == 0 else "CHECK"
    return {
        "source_file": doc.fields.source_file,
        "bill_no": doc.fields.bill_no,
        "check": name,
        "expected": expected,
        "actual": round(actual, 2),
        "variance": variance,
        "status": status,
        "note": note,
    }


def write_workbook(documents: Sequence[Document], path: str | Path) -> Path:
    path = Path(path)
    build_workbook(documents).save(path)
    return path


def workbook_bytes(documents: Sequence[Document]) -> bytes:
    buffer = BytesIO()
    build_workbook(documents).save(buffer)
    return buffer.getvalue()
